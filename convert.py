#!/usr/bin/env python3
from dataclasses import dataclass
from typing import List, Tuple, Dict, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
import xml.etree.ElementTree as ET
from xml.dom import minidom
import xmlschema
import sys


def elem(tag, attrib=None, text=None, children=None, **extra) -> ET.Element:
    e = ET.Element(tag, (attrib or {}), **extra)
    e.text = text
    if children is not None:
        e.extend([c for c in children if c is not None])
    return e


@dataclass
class BatchDataSet:
    schema_version: str
    quarter: str
    year: str
    reporting_state: str
    submission_type: str

    def to_xml(self) -> ET.Element:
        return elem(
            "BatchDataSet",
            attrib={
                "SchemaVersion": str(acc.batch_data_set.schema_version),
                "Quarter": str(acc.batch_data_set.quarter),
                "Year": str(acc.batch_data_set.year),
                "ReportingState": str(acc.batch_data_set.reporting_state),
                "SubmissionType": str(acc.batch_data_set.submission_type),
            },
        )


@dataclass
class Contact:
    first_name: str
    middle_name: Optional[str]
    last_name: str
    name_suffix: Optional[str]
    email_address: str
    address: str
    address_2: Optional[str]
    city: str
    state: str
    postal_code: str
    postal_country_code: str
    phone_country: Optional[str]
    phone_area: str
    phone_prefix: str
    phone_line: str
    phone_extention: Optional[str]
    fax: str

    def to_xml(self) -> ET.Element:
        return elem(
            "Contact",
            children=[
                elem("FirstName", text=self.first_name),
                elem("MiddleName", text=self.middle_name) if self.middle_name else None,
                elem("LastName", text=self.last_name),
                elem("NameSuffix", text=self.name_suffix) if self.name_suffix else None,
                elem("EmailAddress", text=self.email_address),
                elem(
                    "ContactAddress",
                    children=[
                        elem("Address", text=self.address),
                        elem("Address2", text=self.address_2)
                        if self.address_2
                        else None,
                        elem("City", text=self.city),
                        elem("StateCode", text=self.state),
                        elem("PostalCode", text=self.postal_code),
                        elem("CountryCode", text=self.postal_country_code),
                    ],
                ),
                elem(
                    "PhoneNumber",
                    children=[
                        elem("CountryCode", text=self.phone_country)
                        if self.phone_country
                        else None,
                        elem("AreaCode", text=self.phone_area),
                        elem("Prefix", text=self.phone_prefix),
                        elem("Line", text=self.phone_line),
                        elem("Extension", text=self.phone_extention)
                        if self.phone_extention
                        else None,
                    ],
                ),
                elem("Fax", text=self.fax),
            ],
        )


@dataclass
class Insurer:
    naic_number: str
    name: str
    contact: Contact

    def to_xml(self) -> ET.Element:
        return elem(
            "Insurer",
            children=[
                elem("NAICNumber", text=self.naic_number),
                elem("Name", text=self.name),
                self.contact.to_xml(),
            ],
        )


@dataclass(eq=True, frozen=True)
class Broker:
    license_number: str
    first_name: Optional[str]
    last_name: Optional[str]

    def to_xml(self) -> ET.Element:
        return elem(
            "Broker",
            children=[
                elem("LicenseNumber", text=self.license_number),
                elem("FirstName", text=self.first_name) if self.first_name else None,
                elem("LastName", text=self.last_name) if self.last_name else None,
            ],
        )


@dataclass(eq=True, frozen=True)
class Policy:
    number: str
    effective_date: str
    expiration_date: str
    insured_name: str

    def to_xml(self) -> ET.Element:
        return elem(
            "Policy",
            children=[
                elem("PolicyNumber", text=self.number),
                elem("EffectiveDate", text=self.effective_date),
                elem("ExpirationDate", text=self.expiration_date),
                elem("InsuredName", text=self.insured_name),
            ],
        )


@dataclass(eq=True, frozen=True)
class Transaction:
    effective_date: str
    premium: str

    def to_xml(self) -> ET.Element:
        return elem(
            "Transaction",
            children=[
                elem("EffectiveDate", text=self.effective_date),
                elem("Premium", text=str(round(float(self.premium), 2))),
            ],
        )


def max_str(s: str, max_len: int) -> str:
    """Returns stripped s w/ given max len (accounting for special chars)"""
    s = s.strip()
    s = s[0 : min(len(s), max_len)]
    further_reductions = (
        s.count("&") * 4  # &amp;
        + s.count("'") * 5  # &apos;
        + s.count('"') * 5  # &quot;
        + s.count("<") * 3  # &lt;
        + s.count(">") * 3  # &gt;
    )
    s = s[0 : len(s) - further_reductions]
    return s


@dataclass
class Accumulator:
    batch_data_set: BatchDataSet
    insurer: Insurer
    data: Dict[Broker, Dict[Policy, List[Transaction]]]

    @property
    def num_transactions(self):
        return sum(
            len(transactions)
            for policy_dict in self.data.values()
            for transactions in policy_dict.values()
        )

    @staticmethod
    def parse_transaction_row(
        ws: Worksheet, row: int
    ) -> Tuple[Broker, Policy, Transaction]:
        broker = Broker(
            license_number=str(ws[f"AA{row}"].value).strip(),  # min=7, max=7
            first_name=max_str(ws[f"Y{row}"].value, 50),  # min=1,max=50
            last_name=max_str(ws[f"Z{row}"].value, 50),  # min=1,max=50
        )
        policy = Policy(
            number=max_str(ws[f"C{row}"].value, 50),  # min=1,max=50
            effective_date=ws[f"L{row}"].value.date().isoformat(),
            expiration_date=ws[f"N{row}"].value.date().isoformat(),
            insured_name=max_str(ws[f"D{row}"].value, 75),  # min=1,max=75
        )
        transaction = Transaction(
            effective_date=ws[f"M{row}"].value.date().isoformat(),
            premium=round(float(ws[f"R{row}"].value), 2),
        )
        return broker, policy, transaction

    def add_row(self, ws: Worksheet, row: int):
        broker, policy, transaction = self.parse_transaction_row(ws, row)
        broker_policies = self.data.get(broker, {})
        policy_transactions = broker_policies.get(policy, [])
        policy_transactions.append(transaction)
        broker_policies[policy] = policy_transactions
        self.data[broker] = broker_policies


def parse_worksheet(acc: Accumulator, ws: Worksheet):
    for row in range(3, ws.max_row + 1):
        if isinstance(ws[f"A{row}"].value, datetime):
            acc.add_row(ws, row)


def formatted_xml(elem: ET.Element) -> str:
    return minidom.parseString(ET.tostring(elem, "utf-8")).toprettyxml(
        indent="  ", encoding="utf-8"
    )


def generate_xml(acc: Accumulator) -> ET.Element:
    broker_children = []
    broker_counter = 1
    policy_counter = 1
    transaction_counter = 1
    for broker, policy_dict in acc.data.items():
        policy_children = []
        for policy, transactions in policy_dict.items():
            transaction_children = []
            for transaction in transactions:
                transaction_elem = transaction.to_xml()
                transaction_elem.set("Xml_TransactionID", str(transaction_counter))
                transaction_counter += 1
                transaction_children.append(transaction_elem)
            policy_elem = policy.to_xml()
            policy_elem.set("Xml_PolicyID", str(policy_counter))
            policy_counter += 1
            policy_elem.append(elem("Transactions", children=transaction_children))
            policy_children.append(policy_elem)
        broker_elem = broker.to_xml()
        broker_elem.set("Xml_BrokerID", str(broker_counter))
        broker_counter += 1
        broker_elem.append(elem("Policies", children=policy_children))
        broker_children.append(broker_elem)
    batch_data_set_elem = acc.batch_data_set.to_xml()
    batch_data_set_elem.append(acc.insurer.to_xml())
    batch_data_set_elem.append(elem("Brokers", children=broker_children))
    return batch_data_set_elem


if len(sys.argv) <= 1 or sys.argv[1].lower() in ["help", "-h", "-help", "--help"]:
    print("./convert.py <xlsx file> <worksheet name> <year> <quarter>")
    print("Example:")
    print('./convert.py FL_2020_Q4.xlsx "2020 Q3 for xml" 2020 4')
else:
    filename = sys.argv[1]
    wsname = sys.argv[2]
    year = sys.argv[3]
    quarter = sys.argv[4]

    acc = Accumulator(
        batch_data_set=BatchDataSet(
            year=year,
            quarter=quarter,
            schema_version="1.0",
            reporting_state="FL",
            submission_type="INS",
        ),
        insurer=Insurer(
            naic_number="16754",
            name="PALOMAR EXCESS AND SURPLUS INSURANCE COMPANY",
            contact=Contact(
                first_name="Christine",
                middle_name=None,
                last_name="Swann",
                name_suffix=None,
                email_address="Compliance@plmr.com",
                address="7979 Ivanhoe Avenue",
                address_2="Suite 500",
                city="La Jolla",
                state="CA",
                postal_code="92037",
                postal_country_code="USA",
                phone_country=None,
                phone_area="619",
                phone_prefix="567",
                phone_line="4574",
                phone_extention=None,
                fax="6198327327",
            ),
        ),
        data={},
    )
    wb = load_workbook(filename=filename)
    ws = wb[wsname]
    parse_worksheet(acc, ws)
    print(f"Num transactions: {acc.num_transactions}")
    root_elem = generate_xml(acc)
    xml_path = f"FSLSO_{year}_Q{quarter}.xml"
    xsd_path = "FSLSO.Insurer.Schema.xsd"
    with open(xml_path, "wb") as f:
        f.write(formatted_xml(root_elem))
    print("validating xml...")
    try:
        schema = xmlschema.XMLSchema(xsd_path)
        schema.is_valid(xml_path)
        print("IS valid")
    except e:
        print("NOT valid")
        raise e
